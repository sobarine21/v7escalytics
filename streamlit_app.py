import streamlit as st
import google.generativeai as genai
from langdetect import detect
from textblob import TextBlob
from fpdf import FPDF
from io import BytesIO
import concurrent.futures
import json
import docx2txt
from PyPDF2 import PdfReader
import re
import email
from email import policy
from email.parser import BytesParser
import time
from datetime import datetime
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
from gtts import gTTS
import easyocr
import openpyxl

genai.configure(api_key=st.secrets["GOOGLE_API_KEY"])

st.set_page_config(page_title="Advanced Email AI", page_icon="📧", layout="wide")
st.title("📨 Advanced Email AI Analysis & Insights")
st.write("Extract insights, generate professional responses, and analyze emails with AI.")

hide_streamlit_style = """
    <style>
        .css-1r6p8d1 {display: none;} 
        .css-1v3t3fg {display: none;} 
        .css-1r6p8d1 .st-ae {display: none;} 
        header {visibility: hidden;} 
        .css-1tqja98 {visibility: hidden;} 
    </style>
"""
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

features = {
    "sentiment": True,
    "highlights": True,
    "response": True,
    "export": True,
    "tone": True,
    "urgency": False,
    "task_extraction": True,
    "scenario_responses": True,
    "attachment_analysis": True,
    "phishing_detection": True,
    "sensitive_info_detection": True,
    "confidentiality_rating": True,
    "complexity_reduction": True,
    "bias_detection": True,
    "conflict_detection": True,
    "argument_mining": True,
    "metadata_extraction": True,
}

st.sidebar.title("Feature Selection")
for feature in features:
    features[feature] = st.sidebar.checkbox(f"Enable {feature.replace('_', ' ').title()}", value=features[feature])

email_content = st.text_area("📩 Paste your email content here:", height=200)
MAX_EMAIL_LENGTH = 2000

uploaded_file = st.file_uploader("📎 Upload attachment for analysis (optional):", type=["txt", "pdf", "docx", "eml", "msg", "xlsx"])
uploaded_email_file = st.file_uploader("📧 Upload email for thread analysis:", type=["eml", "msg"])

scenario_options = [
    "Customer Complaint",
    "Product Inquiry",
    "Billing Issue",
    "Technical Support Request",
    "General Feedback",
    "Order Status",
    "Shipping Delay",
    "Refund Request",
    "Product Return",
    "Product Exchange",
    "Payment Issue",
    "Subscription Inquiry",
    "Account Recovery",
    "Account Update Request",
    "Cancellation Request",
    "Warranty Claim",
    "Product Defect Report",
    "Delivery Problem",
    "Product Availability",
    "Store Locator",
    "Service Appointment Request",
    "Installation Assistance",
    "Upgrade Request",
    "Compatibility Issue",
    "Product Feature Request",
    "Product Suggestions",
    "Customer Loyalty Inquiry",
    "Discount Inquiry",
    "Coupon Issue",
    "Service Level Agreement (SLA) Issue",
    "Invoice Clarification",
    "Tax Inquiry",
    "Refund Policy Inquiry",
    "Order Modification Request",
    "Credit Card Authorization Issue",
    "Security Inquiry",
    "Privacy Concern",
    "Product Manual Request",
    "Shipping Address Change",
    "Customer Support Availability Inquiry",
    "Live Chat Issue",
    "Email Support Response Inquiry",
    "Online Payment Gateway Issue",
    "E-commerce Website Bug Report",
    "Technical Documentation Request",
    "Mobile App Issue",
    "Software Update Request",
    "Product Recall Notification",
    "Urgent Request"
]

selected_scenario = st.selectbox("Select a scenario for suggested response:", scenario_options)

@st.cache_data(ttl=3600)
def get_ai_response(prompt, email_content):
    try:
        model = genai.GenerativeModel("gemini-1.5-flash")
        response = model.generate_content(prompt + email_content[:MAX_EMAIL_LENGTH])
        return response.text.strip()
    except Exception as e:
        st.error(f"AI Error: {e}")
        return ""

def get_sentiment(email_content):
    return TextBlob(email_content).sentiment.polarity

def get_readability(email_content):
    return round(TextBlob(email_content).sentiment.subjectivity * 10, 2)

def export_pdf(text):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.multi_cell(0, 10, text)
    return pdf.output(dest='S').encode('latin1')

def analyze_phishing_links(email_content):
    phishing_keywords = ["login", "verify", "update account", "account suspended", "urgent action required", "click here"]
    phishing_links = []
    urls = re.findall(r'(https?://\S+)', email_content)
    for url in urls:
        for keyword in phishing_keywords:
            if keyword.lower() in url.lower():
                phishing_links.append(url)
    return phishing_links

def detect_sensitive_information(email_content):
    sensitive_info_patterns = {
        "phone_number": r"(\+?\d{1,2}\s?)?(\(?\d{3}\)?|\d{3})[\s\-]?\d{3}[\s\-]?\d{4}",
        "email_address": r"[\w\.-]+@[\w\.-]+\.\w+",
        "credit_card": r"\b(?:\d[ -]*?){13,16}\b"
    }
    
    sensitive_data = {}
    for key, pattern in sensitive_info_patterns.items():
        matches = re.findall(pattern, email_content)
        if matches:
            sensitive_data[key] = matches
    return sensitive_data

def confidentiality_rating(email_content):
    keywords = ["confidential", "private", "restricted", "not for distribution"]
    rating = sum(1 for keyword in keywords if keyword.lower() in email_content.lower())
    return min(rating, 5)

def analyze_attachment(file):
    try:
        if file.type == "text/plain":
            return file.getvalue().decode("utf-8")
        elif file.type == "application/pdf":
            reader = PdfReader(file)
            text = ""
            for page in reader.pages:
                text += page.extract_text()
            return text
        elif file.type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            return docx2txt.process(file)
        elif file.type in ["message/rfc822", "application/vnd.ms-outlook"]:
            msg = BytesParser(policy=policy.default).parsebytes(file.getvalue())
            return msg.get_body(preferencelist=('plain')).get_content()
        elif file.type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            wb = openpyxl.load_workbook(file)
            summary = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                summary.append(f"Sheet: {sheet}")
                for row in ws.iter_rows(values_only=True):
                    summary.append(", ".join([str(cell) for cell in row if cell is not None]))
            return "\n".join(summary)
        else:
            return "Unsupported file type."
    except Exception as e:
        return f"Error analyzing attachment: {e}"

def extract_email_metadata(email_file):
    try:
        msg = BytesParser(policy=policy.default).parsebytes(email_file.getvalue())
        metadata = {
            "from": msg.get("From"),
            "to": msg.get("To"),
            "subject": msg.get("Subject"),
            "date": msg.get("Date"),
            "message_id": msg.get("Message-ID"),
            "in_reply_to": msg.get("In-Reply-To"),
            "references": msg.get("References"),
            "attachments": [part.get_filename() for part in msg.iter_attachments()]
        }
        return metadata
    except Exception as e:
        return f"Error extracting metadata: {e}"

def progress_bar(duration):
    progress = st.progress(0)
    for i in range(duration):
        time.sleep(1)
        progress.progress((i + 1) / duration)

def visualize_argument_mining(argument_mining):
    arguments = argument_mining.split("\n")
    arguments = [arg for arg in arguments if arg.strip()]
    data = pd.DataFrame({"Argument": arguments})
    plt.figure(figsize=(10, 6))
    sns.countplot(y="Argument", data=data, palette="viridis")
    plt.title("Argument Mining Results")
    plt.xlabel("Count")
    plt.ylabel("Arguments")
    st.pyplot(plt)

def visualize_conflict_detection(conflict_detection):
    conflicts = conflict_detection.split("\n")
    conflicts = [conflict for conflict in conflicts if conflict.strip()]
    data = pd.DataFrame({"Conflict": conflicts})
    plt.figure(figsize=(10, 6))
    sns.countplot(y="Conflict", data=data, palette="Reds")
    plt.title("Conflict Detection Results")
    plt.xlabel("Count")
    plt.ylabel("Conflicts")
    st.pyplot(plt)

def text_to_speech(text):
    tts = gTTS(text)
    tts_bytes = BytesIO()
    tts.write_to_fp(tts_bytes)
    tts_bytes.seek(0)
    return tts_bytes

def copy_to_clipboard(text, element_id):
    st.markdown(f"""
        <textarea id="{element_id}" style="position: absolute; left: -1000px;">{text}</textarea>
        <button onclick="copyText('{element_id}')">Copy to Clipboard</button>
        <script>
            function copyText(elementId) {{
                var copyText = document.getElementById(elementId);
                copyText.select();
                document.execCommand("copy");
            }}
        </script>
    """, unsafe_allow_html=True)

if (email_content or uploaded_file or uploaded_email_file) and st.button("🔍 Generate Insights"):
    try:
        progress_bar(5)

        if uploaded_email_file:
            msg = BytesParser(policy=policy.default).parsebytes(uploaded_email_file.getvalue())
            email_content = msg.get_body(preferencelist=('plain')).get_content()
            metadata = extract_email_metadata(uploaded_email_file)

        detected_lang = detect(email_content)
        if detected_lang != "en":
            st.error("⚠️ Only English language is supported.")
        else:
            with st.spinner("⚡ Processing email insights..."):
                with concurrent.futures.ThreadPoolExecutor() as executor:
                    future_summary = executor.submit(get_ai_response, "Summarize this email concisely:\n\n", email_content) if features["highlights"] else None
                    future_response = executor.submit(get_ai_response, "Generate a professional response to this email:\n\n", email_content) if features["response"] else None
                    future_highlights = executor.submit(get_ai_response, "Highlight key points:\n\n", email_content) if features["highlights"] else None
                    future_tone = executor.submit(get_ai_response, "Detect the tone of this email:\n\n", email_content) if features["tone"] else None
                    future_tasks = executor.submit(get_ai_response, "List actionable tasks:\n\n", email_content) if features["task_extraction"] else None
                    future_complexity_reduction = executor.submit(get_ai_response, "Explain this email in the simplest way possible:\n\n", email_content) if features["complexity_reduction"] else None

                    scenario_prompt = f"Generate a response for a {selected_scenario.lower()}:\n\n"
                    future_scenario_response = executor.submit(get_ai_response, scenario_prompt, email_content) if features["scenario_responses"] else None

                    attachment_text = analyze_attachment(uploaded_file) if uploaded_file and features["attachment_analysis"] else None
                    future_attachment_analysis = executor.submit(get_ai_response, "Analyze this attachment content:\n\n", attachment_text) if attachment_text else None

                    phishing_links = analyze_phishing_links(email_content) if features["phishing_detection"] else []

                    sensitive_info = detect_sensitive_information(email_content) if features["sensitive_info_detection"] else {}

                    confidentiality = confidentiality_rating(email_content) if features["confidentiality_rating"] else 0

                    future_bias_detection = executor.submit(get_ai_response, "Identify potential biases in this email:\n\n", email_content) if features["bias_detection"] else None
                    future_conflict_detection = executor.submit(get_ai_response, "Detect conflicts in this email thread:\n\n", email_content) if features["conflict_detection"] else None
                    future_argument_mining = executor.submit(get_ai_response, "Analyze the arguments presented in this email:\n\n", email_content) if features["argument_mining"] else None

                    if features["metadata_extraction"] and uploaded_email_file:
                        email_metadata = extract_email_metadata(uploaded_email_file)
                    else:
                        email_metadata = None

                    summary = future_summary.result() if future_summary else None
                    response = future_response.result() if future_response else None
                    highlights = future_highlights.result() if future_highlights else None
                    tone = future_tone.result() if future_tone else None
                    tasks = future_tasks.result() if future_tasks else None
                    readability_score = get_readability(email_content)
                    complexity_reduction = future_complexity_reduction.result() if future_complexity_reduction else None
                    scenario_response = future_scenario_response.result() if future_scenario_response else None
                    attachment_analysis = future_attachment_analysis.result() if future_attachment_analysis else None
                    bias_detection = future_bias_detection.result() if future_bias_detection else None
                    conflict_detection = future_conflict_detection.result() if future_conflict_detection else None
                    argument_mining = future_argument_mining.result() if future_argument_mining else None

                col1, col2 = st.columns(2)

                with col1:
                    if summary:
                        with st.expander("📌 Email Summary"):
                            st.write(summary)
                            copy_to_clipboard(summary, "summary_clipboard")

                    if response:
                        with st.expander("✉️ Suggested Response"):
                            st.write(response)
                            copy_to_clipboard(response, "response_clipboard")

                    if highlights:
                        with st.expander("🔑 Key Highlights"):
                            st.write(highlights)
                            copy_to_clipboard(highlights, "highlights_clipboard")

                    if features["sentiment"]:
                        with st.expander("💬 Sentiment Analysis"):
                            sentiment = get_sentiment(email_content)
                            sentiment_label = "Positive" if sentiment > 0 else "Negative" if sentiment < 0 else "Neutral"
                            st.write(f"**Sentiment:** {sentiment_label} (Polarity: {sentiment:.2f})")

                    if tone:
                        with st.expander("🎭 Email Tone"):
                            st.write(tone)

                    if tasks:
                        with st.expander("📝 Actionable Tasks"):
                            st.write(tasks)
                            copy_to_clipboard(tasks, "tasks_clipboard")

                    if complexity_reduction:
                        with st.expander("🔽 Simplified Explanation"):
                            st.write(complexity_reduction)
                            tts_bytes = text_to_speech(complexity_reduction)
                            st.audio(tts_bytes)
                            copy_to_clipboard(complexity_reduction, "complexity_reduction_clipboard")

                    if scenario_response:
                        with st.expander("📜 Scenario-Based Suggested Response"):
                            st.write(f"**{selected_scenario}:**")
                            st.write(scenario_response)
                            copy_to_clipboard(scenario_response, "scenario_response_clipboard")

                    if attachment_analysis:
                        with st.expander("📎 Attachment Analysis"):
                            st.write(attachment_analysis)
                            copy_to_clipboard(attachment_analysis, "attachment_analysis_clipboard")

                    if phishing_links:
                        with st.expander("⚠️ Phishing Links Detected"):
                            st.write(phishing_links)
                            copy_to_clipboard("\n".join(phishing_links), "phishing_links_clipboard")

                    if sensitive_info:
                        with st.expander("⚠️ Sensitive Information Detected"):
                            st.json(sensitive_info)
                            copy_to_clipboard(json.dumps(sensitive_info, indent=4), "sensitive_info_clipboard")

                    if confidentiality:
                        with st.expander("🔐 Confidentiality Rating"):
                            st.write(f"Confidentiality Rating: {confidentiality}/5")

                with col2:
                    if bias_detection:
                        with st.expander("⚖️ Bias Detection"):
                            st.write(bias_detection)

                    if conflict_detection:
                        with st.expander("🚨 Conflict Detection"):
                            st.write(conflict_detection)
                            visualize_conflict_detection(conflict_detection)

                    if argument_mining:
                        with st.expander("💬 Argument Mining"):
                            st.write(argument_mining)
                            visualize_argument_mining(argument_mining)

                    if email_metadata:
                        with st.expander("📅 Email Metadata"):
                            metadata_df = pd.DataFrame(list(email_metadata.items()), columns=["Field", "Value"])
                            st.table(metadata_df)

                if features["export"]:
                    export_data = {
                        "summary": summary,
                        "response": response,
                        "highlights": highlights,
                        "complexity_reduction": complexity_reduction,
                        "scenario_response": scenario_response,
                        "attachment_analysis": attachment_analysis,
                        "phishing_links": phishing_links,
                        "sensitive_info": sensitive_info,
                        "confidentiality": confidentiality,
                        "bias_detection": bias_detection,
                        "conflict_detection": conflict_detection,
                        "argument_mining": argument_mining,
                        "metadata": email_metadata,
                    }
                    export_json = json.dumps(export_data, indent=4)
                    st.download_button("📥 Download JSON", data=export_json, file_name="analysis.json", mime="application/json")

                    pdf_data = export_pdf(json.dumps(export_data, indent=4))
                    st.download_button("📥 Download PDF", data=pdf_data, file_name="analysis.pdf", mime="application/pdf")

                    export_csv = pd.DataFrame.from_dict(export_data, orient='index').to_csv().encode('utf-8')
                    st.download_button("📥 Download CSV", data=export_csv, file_name="analysis.csv", mime="text/csv")

    except Exception as e:
        st.error(f"❌ Error: {e}")

else:
    st.info("✏️ Paste email content and click 'Generate Insights' to begin.")
